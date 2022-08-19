
def templates(input_master,master_worksheet,owner_column):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    import warnings

    warnings.simplefilter("ignore")

    #yellow_highlight_rbg = 'FFFFFF00'


    # DYNAMIC VALUES
    """input_master = "N:\Groups\Operational Risk\Common\Vendor Metadata Gathering\Business Owner Templates\Master File2 Ek.xlsx" 
    master_worksheet = 'Summary AP report' """

    blank_template = "N:\Groups\Operational Risk\Common\Vendor Metadata Gathering\Business Owner Templates\Business_Owners_Blank_new.xlsm"
    blank_worksheet = 'Questionnaire'
    
    owners_needed = 'Templates Needed'

    owners_needed_col = 'A'
    #owner_column = 'Y'
    website_url_col = 'G'
    """param_col = 'Z'
    notes_params = ['Also sent to Yanna for VAO', 'Michele Martell will review', 'Reasigned to Yanna per her request']
    owners_params = ['Martell, Michele'] """

    def main():
        #load excel files
        wb = openpyxl.load_workbook(blank_template, keep_vba = True, read_only=False)
        master_lst = openpyxl.load_workbook(input_master,data_only=True)

        #load worksheets
        ap_report = master_lst.get_sheet_by_name(master_worksheet) 
        target_owners = master_lst.get_sheet_by_name(owners_needed)
        num_owners = target_owners.max_row+1

        # vendorData['Christina Reilly'] = [row1, row2, row5, ...]

        # func make_owner_dict returns dict of owners (key) and list of row numbers (values as lists)
        # func make_owner_template takes owner name and lists (row nums), returns new template for each owner

        # func to get list of cells in target_owners worksheet that are highlighted in yellow
        priority_owners = get_priority_owners(target_owners, num_owners)
        print("Priority Business Owners: ")
        print(priority_owners)

        vendorData = make_owners_rows_dict(ap_report, owner_column)

        for owner in priority_owners:
            #updated_rows=[]
            print(vendorData[owner])
            #updated_rows=[row for row in vendorData[owner] if (len(notes_params)==0 and len(owners_params)==0) or (ap_report[param_col+str(row)].value in notes_params) or (ap_report[owner_column+str(row)].value in owners_params)]
            #make_owner_template(owner, ap_report, wb, updated_rows)
            make_owner_template(owner, ap_report, wb, vendorData[owner])
            print()

    # return array of owner names that are highlighted
    def get_priority_owners(target_owners, num_owners):
        highlighted_names = []

        #add 1 to num_owners for header
        for row in range(2, num_owners+1):
            cell = target_owners[owners_needed_col+str(row)]
            #if cell.fill.fgColor.rgb == yellow_highlight_rbg:
            if cell.fill.fgColor.rgb != '00000000':
                highlighted_names.append(cell.value)
        return highlighted_names

    #func makes dict of ALL owners in ap_report worksheet
    def make_owners_rows_dict(ap_report, owner_column):
        vendorData = {}

        for row in range(3, ap_report.max_row+1):
            #each row in the spreadsheet has data for one vendor
            owner = ap_report[owner_column + str(row)].value
            
            if owner == None or owner == 0:
                continue
            elif owner not in vendorData:
                vendorData.setdefault(owner, [])
            vendorData[owner].append(row)

        return vendorData


    # Make template files
    def make_owner_template(owner, ap_report, wb, updated_rows):
        #print(owner, updated_rows)
        worksheet = wb.get_sheet_by_name(blank_worksheet)

        #data validations
        level2 = DataValidation(type='list', formula1="=OFFSET(Vendors!$L$23,1,MATCH(J5,Vendors!$L$23:$X$23,0)-1,COUNTA(OFFSET(Vendors!$L$23,1,MATCH(J5,Vendors!$L$23:$X$23,0)-1,20,1)),1)")
        level3 = DataValidation(type='list', formula1="=OFFSET(Vendors!$M$3,1,MATCH(K5,Vendors!$M$3:$T$3,0)-1,COUNTA(OFFSET(Vendors!$M$3,1,MATCH(K5,Vendors!$M$3:$T$3,0)-1,20,1))-1,1)")

        worksheet.add_data_validation(level2)
        worksheet.add_data_validation(level3)
        level2.add('K5:K100000')
        level3.add('L5:L100000')

        #clear cell values from columns A (Alternate Supplier Name) through G (websites)
        for row in worksheet['A5':'G1000']:
            for cell in row:
                cell.value = None	


        #make business templates
        #if if_sent_to_greg_anthony == 0 or if_sent_to_greg_anthony == None:
        row_curr = 5
        for row in updated_rows:
            
            # notes_col = ap_report[param_col+str(updated_rows[i])].value
            # questionnaire_owner2_col = ap_report[owner_column+str(updated_rows[i])].value
            
            #if_sent_to_greg_anthony = ap_report['L'+str(updated_rows[i])].value
        
            #if (len(notes_params) == 0 and len(owners_params) == 0) or (notes_col in notes_params) or (questionnaire_owner2_col in owners_params):
            #updated_rownums.append(updated_rows[i])
            supplier_id = ap_report['A'+str(row)].value
            supplier_name = ap_report['B'+str(row)].value
            business_owner = owner
            alternate_supplier_name = ap_report['C'+str(row)].value
            supplier_category = ap_report['D'+str(row)].value
            website_url = ap_report['W'+str(row)].value

            worksheet['A'+str(row_curr)] = alternate_supplier_name
            worksheet['B'+str(row_curr)] = supplier_category
            worksheet['C'+str(row_curr)] = supplier_id
            worksheet['D'+str(row_curr)] = supplier_name
            worksheet['E'+str(row_curr)] = business_owner
            #worksheet[website_url_col+str(row_curr)] = website_url
            row_curr += 1

        wb.save('N:/Groups/Operational Risk/Common/Vendor Metadata Gathering/Business Owner Templates/EXAMPLES/OUT_Business Owners Questionnaire_'+str(owner)+'.xlsm')               
        #wb.save('N:/Groups/Operational Risk/Common/Vendor Metadata Gathering/Business Owner Templates/Business Owners Questionnaire2_'+str(owner)+'.xlsm')
    main()

#if notes_col == 'Also sent to Yanna for VAO' or notes_col == 'Michele Martell will review' or notes_col == 'Reasigned to Yanna per her request' or questionnaire_owner2_col == 'Martell, Michele'
#worksheet[website_url_col+str(row_curr)] = website_url

#print('All owner and row nums:')
#for owner, row in vendorData.items():
#	print(owner, row)

#make_owner_template('Christina Reilly', vendorData['Christina Reilly'], ap_report, wb)

#for owner in owners_lst[3: questionnaire_owner_lst.max_row]:
#	#vendor_count = questionnaire_owner_lst['B' + str(owner.row)].value
#	make_owner_template(owner, vendorData[owner], ap_report, wb)