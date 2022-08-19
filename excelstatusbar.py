# -*- coding: utf-8 -*-


# LOAD EXCEL WBs AND SHEETS
    # master
#input_master = "N:/Groups/Operational Risk/Common/Vendor Metadata Gathering/Business Owner Templates/Master File2 Ek - Copy.xlsx"
#master_worksheet = 'Summary AP report'

    # team-specific (Questionnaires)
#input_sent_template = "N:\Groups\Operational Risk\Common\Vendor Metadata Gathering\Business Owner Templates\Owner Templates\GK_Business Owners Questionnaire_Technology_first.xlsm"
#input_worksheet = 'Questionnaire'


#team = 'Technology'


def generate_status(input_master,master_worksheet,input_sent_template,input_worksheet,team, input_flagged_col, last_row):
    # MODULES
    import openpyxl
    import warnings
    import sys
    #from generatetemplates import templates

    #print(sys.path)
    sys.path.append('U:/pythonproj/')
    sys.path.append('N:\\Groups\\Operational Risk\\Common\\Vendor Metadata Gathering\\Business Owner Templates')
    from processflagged import process_flagged


    warnings.simplefilter("ignore")

    # PARAMETER VARIABLES
    #yellow_rgb = 'FFFFFF00'
    #red_rgb = 'FFFF0000'

    # OUTPUT PATH
    orig_stdout = sys.stdout
    output_file = open(f'N:\Groups\Operational Risk\Common\Vendor Metadata Gathering\Business Owner Templates\EXAMPLES\\{team}_08182022.txt', 'w')
    #output_file = open(f'N:\Groups\Operational Risk\Common\Vendor Metadata Gathering\Business Owner Templates\Questionnaires Completion Progress (Stats)\\{team}_08182022.txt', 'w')
    sys.stdout = output_file

    next_deadline = 'July 29th'

    # DYNAMIC VARIABLES
    start_col = 8 #H (right after "Vendor Website" column)
    end_col = 28 #AB
    start_row = 5
    percent_completed_threshold = 100
    master_col_dict_key = 'A'

    # if_sent is manually completed in the excel file and refreshed as needed into db
    # if_remove is 'Yes' if owner flags it
    # if_reassigned is 'Yes' if there is a note in the column
    # if_completed is 'Yes' if completion status is 100%
    #if_sent = 'Z'
    #if_remove = 'AA'
    #if_reassigned = 'AB'
    #if_completed = 'AC'



    def main(input_flagged_col):

        # ====================== LOAD EXCEL FILES =======================================================

        #load master
        master_lst = openpyxl.load_workbook(str(input_master))
        ap_report = master_lst.get_sheet_by_name(str(master_worksheet))
            
        #load team-specific
        wb = openpyxl.load_workbook(str(input_sent_template), keep_vba = True, read_only=False)
        questionnaire_sheet = wb.get_sheet_by_name(str(input_worksheet))

        #vendor_rows_dict = templates.make_owners_rows_dict(ap_report, )

        # =============================================================================
        
        completed_supplier_ids, flagged_supplier_ids, flagged_rows = check_status(questionnaire_sheet, ap_report)
        #print()
        #print(flagged_rows)
        
        # 'process_flagged' function updates the master spreadsheet based on the flagged rows (vendors) and reassignments/notes
        master_lst = process_flagged(questionnaire_sheet, flagged_rows, master_lst, team, master_col_dict_key, input_flagged_col)

        master_lst.save(input_master)
        

        #print(nonblanks)
        return completed_supplier_ids, flagged_supplier_ids
        
    
    # Print stats and progress bars for each vendor as .txt file
    def check_status(sheet,ap_report):
        
        # identified all vendors that are completed to the previously specified 'percent_completed_threshold'
        # also return the rows and supplier IDs of the vendors flagged by Owner
        row = start_row
        completed = 0
        flagged_rows = []
        flagged_supplier_ids = []
        incomplete_rows = []
        completed_cols = []
        completed_supplier_ids=[]
            
        while sheet.cell(row=row, column=6).value != next_deadline and row <= last_row:
            cells_per_row = 0
            filled_cells = 0
            supplier_id = sheet.cell(row=row, column=3).value
            #supplier_name = sheet.cell(row=row, column=4).value       
            
            for col in range(start_col, end_col):
                
                #exceptions (skip over columns K, L, N, O in each row -> not necessary)
                if col == 11 or col == 12 or col == 14 or col == 15 or col == 28:
                    continue
            
                #check if any cell is highlighted red or yellow: skip after incrementing reassigned
                if sheet.cell(row=row, column=col).fill.fgColor.rgb != '00000000':
                #if sheet.cell(row=row, column=col).fill.fgColor.rgb == yellow_rgb or sheet.cell(row=row, column=col).fill.fgColor.rgb == red_rgb:
                    flagged_rows.append(row)
                    flagged_supplier_ids.append(supplier_id)
                    #completed+=1
                    break
                    
                #if value is not NULL count cell as being "completed"
                if sheet.cell(row=row, column=col).value not in ['None', '', None]:
                    filled_cells += 1
                cells_per_row += 1
                
            #get percent completed across each row (for each vendor)
            if cells_per_row > 0:
                completed_percent = round((filled_cells / cells_per_row)*100,2) 
                
                #check if 100% completed
                if completed_percent >= percent_completed_threshold:
                    completed+=1
                    ap_report.cell(row=row-1,column=29).value = 'Yes'
                else:
                    incomplete_rows.append(supplier_id)
                
                #append completed % and supplier ids to their respective lists
                #to determine which suppliers are done
                completed_cols.append(completed_percent)
                completed_supplier_ids.append(supplier_id)
                    
                print(f'{row}, {supplier_id}')
                progress(supplier_id, int(completed_percent))
            
            row += 1
        
        #get the total nums and stats
        flagged_num = len(flagged_rows)
        suppliers_num_deadline = row-start_row
        total_percent_completed = round((completed / suppliers_num_deadline)*100,2)
        
        print('***********************STATS**********************************')
        
        print(f'Total number Vendors (up to {next_deadline} deadline): {suppliers_num_deadline}')
        print(f'Total number of Vendors Completed: {completed}')
        print(f'Percentage of Vendors Completed up to {next_deadline} deadline: {total_percent_completed}%')
        print(f'Number of Vendors Flagged: {flagged_num}')
        #print(f'Suppliers not yet completed (up to {next_deadline} deadline): {incomplete_rows}')
        
        #return completed_cols, completed_supplier_ids
        
        #return flagged_rows

        return completed_supplier_ids, flagged_supplier_ids, flagged_rows



    # Print out progress bars for each vendor in Input File
    def progress(supplier_id, percent, width=40):
        left = width * percent // 100
        right = width - left
        
        tags = "#" * left
        spaces = " " * right
        percents = f"{percent:.0f}%"
        
        print("\r[", tags, spaces, "]", percents, sep="", end="", flush=True)
        print()
        print()
        
                
    completed_supplier_ids, flagged_supplier_ids = main(input_flagged_col)

    sys.stdout = orig_stdout
    output_file.close()

    return completed_supplier_ids, flagged_supplier_ids



# =============================================================================
# def plot_status(x,y):
#     #t = np.arange(0.0, 2.0, 0.01)
#     #s = 1 + np.sin(2*np.pi*t)
#     plt.bar(x,y)
# 
#     # setting x-label as pen sold
#     plt.xlabel("pen sold")
#  
#     # setting y_label as price
#     plt.ylabel("price")  
#     plt.title(" Vertical bar graph")
#     plt.show()
# =============================================================================
